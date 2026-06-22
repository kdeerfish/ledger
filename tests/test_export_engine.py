#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出引擎测试
"""

import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import ledger_modules.db as db_module
import ledger_modules.transactions as tx_module
import ledger_modules.export_engine as export_engine


@pytest.fixture
def export_engine_db(temp_db):
    """让导出引擎使用有示例数据的临时数据库（中文类型名）"""
    export_engine.DB_PATH = temp_db
    # 用中文类型名创建测试数据
    tx_module.DB_PATH = temp_db
    tx_module.add_transaction("支出", 100.0, "食品", "零食", "微信", "项目A", "本人", "拼多多", "零食", "2026-06-15 10:00:00")
    tx_module.add_transaction("支出", 200.0, "交通", "打车", "支付宝", "项目A", "本人", "滴滴", "打车", "2026-06-14 14:30:00")
    tx_module.add_transaction("收入", 5000.0, "工资", "", "银行", "", "本人", "", "6月工资", "2026-06-10 09:00:00")
    tx_module.add_transaction("支出", 30.0, "食品", "水果", "微信", "", "成员A", "水果店", "水果", "2026-06-12 18:00:00")
    tx_module.add_transaction("支出", 150.0, "餐饮", "午餐", "xxx信用卡", "个人项目", "本人", "食堂", "午餐", "2026-06-13 12:00:00")
    return temp_db


# ─── 数据查询 ──────────────────────────────────────────────

class TestGetExportData:
    def test_basic(self, export_engine_db):
        data = export_engine.get_export_data()
        assert data['count'] == 5
        assert len(data['transactions']) == 5
        assert data['summary']['income'] == 5000.0
        assert data['summary']['expense'] == 480.0

    def test_filter_by_category(self, export_engine_db):
        data = export_engine.get_export_data(category='食品')
        assert data['count'] == 2
        assert all(t['category'] == '食品' for t in data['transactions'])

    def test_filter_by_type(self, export_engine_db):
        data = export_engine.get_export_data(type_='收入')
        assert data['count'] == 1
        assert data['transactions'][0]['type'] == '收入'

    def test_by_month(self, export_engine_db):
        data = export_engine.get_export_data()
        assert len(data['by_month']) == 1
        assert data['by_month'][0]['month'] == '2026-06'

    def test_by_category(self, export_engine_db):
        data = export_engine.get_export_data()
        cats = [c['category'] for c in data['by_category']]
        assert '食品' in cats
        assert '工资' in cats

    def test_by_account(self, export_engine_db):
        data = export_engine.get_export_data()
        accs = [a['account'] for a in data['by_account']]
        assert '微信' in accs


# ─── Excel 导出 ──────────────────────────────────────────────

class TestExportExcel:
    def test_basic(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.xlsx')
        result = export_engine.export_excel(data, output)
        assert os.path.exists(result)
        assert os.path.getsize(result) > 0

    def test_multiple_sheets(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.xlsx')
        export_engine.export_excel(data, output, sheets=['明细', '月度汇总', '分类统计', '账户统计'])

        import openpyxl
        wb = openpyxl.load_workbook(output)
        assert '交易明细' in wb.sheetnames
        assert '月度汇总' in wb.sheetnames
        assert '分类统计' in wb.sheetnames
        assert '账户统计' in wb.sheetnames

    def test_single_sheet(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.xlsx')
        export_engine.export_excel(data, output, sheets=['明细'])

        import openpyxl
        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 1
        assert '交易明细' in wb.sheetnames


# ─── CSV 导出 ──────────────────────────────────────────────

class TestExportCsv:
    def test_import_compatible(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.csv')
        export_engine.export_csv(data, output, import_compatible=True)

        with open(output, 'r', encoding='utf-8-sig') as f:
            header_line = f.readline().strip()
        headers = header_line.split(',')
        # 应该是导入兼容格式
        assert '交易类型' in headers
        assert '日期' in headers
        assert '金额' in headers

    def test_standard_format(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.csv')
        export_engine.export_csv(data, output, import_compatible=False)

        with open(output, 'r', encoding='utf-8-sig') as f:
            header_line = f.readline().strip()
        headers = header_line.split(',')
        assert 'ID' in headers
        assert '类型' in headers


# ─── JSON 导出 ──────────────────────────────────────────────

class TestExportJson:
    def test_basic(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.json')
        export_engine.export_json(data, output)

        with open(output, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        assert 'summary' in json_data
        assert 'transactions' in json_data
        assert json_data['count'] == 5
        assert len(json_data['transactions']) == 5

    def test_has_tags_field(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.json')
        export_engine.export_json(data, output)

        with open(output, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        assert 'tags' in json_data['transactions'][0]


# ─── PDF 导出 ──────────────────────────────────────────────

class TestExportPdf:
    def test_basic(self, export_engine_db, tmp_path):
        data = export_engine.get_export_data()
        output = str(tmp_path / 'test.pdf')
        result = export_engine.export_pdf(data, output, title='测试报告')
        assert os.path.exists(result)
        assert os.path.getsize(result) > 0


# ─── 预览 ──────────────────────────────────────────────

class TestExportPreview:
    def test_basic(self, export_engine_db):
        result = export_engine.get_export_preview()
        assert result['count'] == 5
        assert result['income'] == 5000.0
        assert result['expense'] == 480.0

    def test_with_filter(self, export_engine_db):
        result = export_engine.get_export_preview(type_='收入')
        assert result['count'] == 1
