"""
增强导出引擎

支持格式：
- Excel (.xlsx): 多 Sheet（明细+月度汇总+分类统计+账户统计）
- PDF: 月度报告（表格+图表）
- CSV: 修复兼容性（列名与导入一致）
- JSON: 标准格式
"""

import csv
import io
import json
import os
import sqlite3
from datetime import datetime

from .db import DB_PATH
from .transaction_types import (
    TYPE_ALIASES,
    EXPENSE,
    INCOME,
    RECONCILIATION,
    TRANSFER,
    is_stat_expense,
    is_stat_income,
    is_transfer,
    normalize_raw_type,
)


# ─── 数据查询 ──────────────────────────────────────────────

def get_export_data(start_date=None, end_date=None, category=None,
                    account=None, type_=None, tag_ids=None):
    """
    查询导出数据，返回结构化结果

    返回:
    {
        "transactions": [...],
        "summary": { "income": ..., "expense": ..., "balance": ... },
        "by_category": [...],
        "by_account": [...],
        "by_month": [...],
        "by_tag": [...],
        "count": N
    }
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # 构建筛选条件
    where_clauses = ["t.is_deleted = 0"]
    params = []

    if start_date:
        where_clauses.append("t.trans_date >= ?")
        params.append(start_date)
    if end_date:
        where_clauses.append("t.trans_date <= ?")
        params.append(end_date + ' 23:59:59' if len(end_date) == 10 else end_date)
    if category:
        where_clauses.append("(t.category = ? OR t.subcategory = ?)")
        params.extend([category, category])
    if account:
        where_clauses.append("t.account = ?")
        params.append(account)
    if type_:
        where_clauses.append("t.type = ?")
        params.append(type_)
    if tag_ids:
        placeholders = ','.join('?' * len(tag_ids))
        where_clauses.append(f"""t.id IN (
            SELECT transaction_id FROM transaction_tags WHERE tag_id IN ({placeholders})
        )""")
        params.extend(tag_ids)

    where_sql = " AND ".join(where_clauses)

    # 查询交易明细（含标签）
    c.execute(f'''SELECT t.id, t.trans_date, t.type, t.amount, t.category,
                         t.subcategory, t.account, t.project, t.member,
                         t.merchant, t.note, t.extra_data, t.batch_id
                  FROM transactions t
                  WHERE {where_sql}
                  ORDER BY t.trans_date ASC''', params)
    rows = c.fetchall()

    transactions = []
    for row in rows:
        tx_id = row['id']
        # 获取标签
        c.execute('''SELECT name FROM tags
                     JOIN transaction_tags ON tags.id = transaction_tags.tag_id
                     WHERE transaction_tags.transaction_id = ?''', (tx_id,))
        tags = [r[0] for r in c.fetchall()]

        transactions.append({
            'id': tx_id,
            'date': row['trans_date'],
            'type': row['type'],
            'amount': row['amount'],
            'category': row['category'] or '',
            'subcategory': row['subcategory'] or '',
            'account': row['account'] or '',
            'project': row['project'] or '',
            'member': row['member'] or '',
            'merchant': row['merchant'] or '',
            'note': row['note'] or '',
            'extra_data': row['extra_data'] or '',
            'batch_id': row['batch_id'],
            'tags': tags,
        })

    # 月度汇总
    c.execute(f'''SELECT strftime('%Y-%m', t.trans_date) as month, t.type,
                         SUM(t.amount), COUNT(*)
                  FROM transactions t
                  WHERE {where_sql}
                  GROUP BY month, t.type
                  ORDER BY month''', params)
    month_rows = c.fetchall()

    by_month = {}
    for row in month_rows:
        month = row[0]
        if month not in by_month:
            by_month[month] = {'month': month, 'income': 0, 'expense': 0, 'count': 0}
        if is_stat_income(row[1]):
            by_month[month]['income'] = row[2]
        elif is_stat_expense(row[1]):
            by_month[month]['expense'] = row[2]
        by_month[month]['count'] += row[3]

    # 分类统计
    c.execute(f'''SELECT t.category, t.type, SUM(t.amount), COUNT(*)
                  FROM transactions t
                  WHERE {where_sql} AND t.category != ''
                  GROUP BY t.category, t.type
                  ORDER BY SUM(t.amount) DESC''', params)
    cat_rows = c.fetchall()

    by_category = {}
    for row in cat_rows:
        cat = row[0]
        if cat not in by_category:
            by_category[cat] = {'category': cat, 'income': 0, 'expense': 0, 'count': 0}
        if is_stat_income(row[1]):
            by_category[cat]['income'] = row[2]
        elif is_stat_expense(row[1]):
            by_category[cat]['expense'] = row[2]
        by_category[cat]['count'] += row[3]

    # 账户统计
    c.execute(f'''SELECT t.account, t.type, SUM(t.amount), COUNT(*)
                  FROM transactions t
                  WHERE {where_sql} AND t.account != ''
                  GROUP BY t.account, t.type
                  ORDER BY SUM(t.amount) DESC''', params)
    acc_rows = c.fetchall()

    by_account = {}
    for row in acc_rows:
        acc = row[0]
        if acc not in by_account:
            by_account[acc] = {'account': acc, 'income': 0, 'expense': 0, 'count': 0}
        if is_stat_income(row[1]):
            by_account[acc]['income'] = row[2]
        elif is_stat_expense(row[1]):
            by_account[acc]['expense'] = row[2]
        by_account[acc]['count'] += row[3]

    # 标签统计
    c.execute(f'''SELECT tg.name, t.type, SUM(t.amount), COUNT(*)
                  FROM transactions t
                  JOIN transaction_tags tt ON t.id = tt.transaction_id
                  JOIN tags tg ON tt.tag_id = tg.id
                  WHERE {where_sql}
                  GROUP BY tg.name, t.type
                  ORDER BY SUM(t.amount) DESC''', params)
    tag_rows = c.fetchall()

    by_tag = {}
    for row in tag_rows:
        tag = row[0]
        if tag not in by_tag:
            by_tag[tag] = {'tag': tag, 'income': 0, 'expense': 0, 'count': 0}
        if is_stat_income(row[1]):
            by_tag[tag]['income'] = row[2]
        elif is_stat_expense(row[1]):
            by_tag[tag]['expense'] = row[2]
        by_tag[tag]['count'] += row[3]

    conn.close()

    # 总计
    total_income = sum(t['amount'] for t in transactions if is_stat_income(t['type']))
    total_expense = sum(t['amount'] for t in transactions if is_stat_expense(t['type']))
    total_transfer = sum(t['amount'] for t in transactions if is_transfer(t['type']))

    return {
        'transactions': transactions,
        'summary': {
            'income': total_income,
            'expense': total_expense,
            'transfer': total_transfer,
            'balance': total_income - total_expense,
        },
        'by_month': sorted(by_month.values(), key=lambda x: x['month']),
        'by_category': sorted(by_category.values(), key=lambda x: x['expense'], reverse=True),
        'by_account': sorted(by_account.values(), key=lambda x: x['expense'], reverse=True),
        'by_tag': sorted(by_tag.values(), key=lambda x: x['expense'], reverse=True),
        'count': len(transactions),
    }


# ─── Excel 导出 ──────────────────────────────────────────────

def export_excel(data, output_path, sheets=None):
    """
    导出 Excel 文件

    参数:
        data: get_export_data() 的返回值
        output_path: 输出文件路径
        sheets: 要包含的 Sheet 列表，默认 ['明细', '月度汇总', '分类统计', '账户统计']
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers

    if sheets is None:
        sheets = ['明细', '月度汇总', '分类统计', '账户统计']

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    money_format = '#,##0.00'
    header_align = Alignment(horizontal='center', vertical='center')

    def style_header(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ''
                    # 中文字符算2个宽度
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, char_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Sheet 1: 交易明细
    if '明细' in sheets:
        ws = wb.active
        ws.title = '交易明细'
        headers = ['ID', '日期', '类型', '金额', '类别', '子类别', '账户',
                    '商家', '成员', '项目', '备注', '标签', '其他信息']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        for i, tx in enumerate(data['transactions'], 2):
            ws.cell(row=i, column=1, value=tx['id'])
            ws.cell(row=i, column=2, value=tx['date'])
            ws.cell(row=i, column=3, value=tx['type'])
            cell = ws.cell(row=i, column=4, value=tx['amount'])
            cell.number_format = money_format
            ws.cell(row=i, column=5, value=tx['category'])
            ws.cell(row=i, column=6, value=tx['subcategory'])
            ws.cell(row=i, column=7, value=tx['account'])
            ws.cell(row=i, column=8, value=tx['merchant'])
            ws.cell(row=i, column=9, value=tx['member'])
            ws.cell(row=i, column=10, value=tx['project'])
            ws.cell(row=i, column=11, value=tx['note'])
            ws.cell(row=i, column=12, value=', '.join(tx.get('tags', [])))
            ws.cell(row=i, column=13, value=tx.get('extra_data', ''))

        auto_width(ws)

    # Sheet 2: 月度汇总
    if '月度汇总' in sheets:
        ws2 = wb.create_sheet('月度汇总')
        headers = ['月份', '收入', '支出', '结余', '笔数']
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h)
        style_header(ws2, 1, len(headers))

        for i, m in enumerate(data['by_month'], 2):
            ws2.cell(row=i, column=1, value=m['month'])
            cell = ws2.cell(row=i, column=2, value=m['income'])
            cell.number_format = money_format
            cell = ws2.cell(row=i, column=3, value=m['expense'])
            cell.number_format = money_format
            cell = ws2.cell(row=i, column=4, value=m['income'] - m['expense'])
            cell.number_format = money_format
            ws2.cell(row=i, column=5, value=m['count'])

        auto_width(ws2)

    # Sheet 3: 分类统计
    if '分类统计' in sheets:
        ws3 = wb.create_sheet('分类统计')
        total_expense = data['summary']['expense'] or 1
        headers = ['类别', '收入合计', '支出合计', '笔数', '支出占比(%)']
        for col, h in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=h)
        style_header(ws3, 1, len(headers))

        for i, cat in enumerate(data['by_category'], 2):
            ws3.cell(row=i, column=1, value=cat['category'])
            cell = ws3.cell(row=i, column=2, value=cat['income'])
            cell.number_format = money_format
            cell = ws3.cell(row=i, column=3, value=cat['expense'])
            cell.number_format = money_format
            ws3.cell(row=i, column=4, value=cat['count'])
            pct = round(cat['expense'] / total_expense * 100, 1) if total_expense else 0
            ws3.cell(row=i, column=5, value=pct)

        auto_width(ws3)

    # Sheet 4: 账户统计
    if '账户统计' in sheets:
        ws4 = wb.create_sheet('账户统计')
        headers = ['账户', '收入合计', '支出合计', '笔数']
        for col, h in enumerate(headers, 1):
            ws4.cell(row=1, column=col, value=h)
        style_header(ws4, 1, len(headers))

        for i, acc in enumerate(data['by_account'], 2):
            ws4.cell(row=i, column=1, value=acc['account'])
            cell = ws4.cell(row=i, column=2, value=acc['income'])
            cell.number_format = money_format
            cell = ws4.cell(row=i, column=3, value=acc['expense'])
            cell.number_format = money_format
            ws4.cell(row=i, column=4, value=acc['count'])

        auto_width(ws4)

    wb.save(output_path)
    return output_path


# ─── CSV 导出（与导入兼容） ──────────────────────────────────────

def export_csv(data, output_path, import_compatible=True):
    """
    导出 CSV 文件

    参数:
        import_compatible=True: 列名与导入格式一致（可直接重新导入）
        import_compatible=False: 使用原有导出格式
    """
    if import_compatible:
        # 与导入格式兼容的列名
        headers = ['交易类型', '日期', '金额', '类别', '子类别', '账户',
                    '项目', '成员', '商家', '备注']
    else:
        # 原有导出格式
        headers = ['ID', '日期', '类型', '金额', '类别', '子类别', '账户',
                    '项目', '成员', '商家', '备注']

    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for tx in data['transactions']:
            if import_compatible:
                # 日期格式转回 YYYY/MM/DD HH:MM
                date_str = tx['date']
                if date_str:
                    try:
                        dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        date_str = dt.strftime('%Y/%m/%d %H:%M')
                    except ValueError:
                        pass
                writer.writerow([
                    tx['type'], date_str, tx['amount'],
                    tx['category'], tx['subcategory'], tx['account'],
                    tx['project'], tx['member'], tx['merchant'], tx['note']
                ])
            else:
                writer.writerow([
                    tx['id'], tx['date'], tx['type'], tx['amount'],
                    tx['category'], tx['subcategory'], tx['account'],
                    tx['project'], tx['member'], tx['merchant'], tx['note']
                ])

    return output_path


# ─── JSON 导出 ──────────────────────────────────────────────

def export_json(data, output_path):
    """导出 JSON 文件"""
    export_data = {
        'summary': data['summary'],
        'count': data['count'],
        'transactions': [],
        'by_month': data['by_month'],
        'by_category': data['by_category'],
        'by_account': data['by_account'],
    }

    for tx in data['transactions']:
        export_data['transactions'].append({
            'id': tx['id'],
            'date': tx['date'],
            'type': tx['type'],
            'amount': tx['amount'],
            'category': tx['category'],
            'subcategory': tx['subcategory'],
            'account': tx['account'],
            'project': tx['project'],
            'member': tx['member'],
            'merchant': tx['merchant'],
            'note': tx['note'],
            'tags': tx.get('tags', []),
            'extra_data': tx.get('extra_data', ''),
        })

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)

    return output_path


# ─── PDF 导出 ──────────────────────────────────────────────

def export_pdf(data, output_path, title=None):
    """
    导出 PDF 报告

    包含：总览 + 分类统计表格 + 明细表格
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    if not title:
        title = '收支报告'

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    # 中文字体支持
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        # 尝试注册中文字体
        font_paths = [
            'C:/Windows/Fonts/msyh.ttc',      # 微软雅黑
            'C:/Windows/Fonts/simsun.ttc',     # 宋体
            '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',  # Linux
        ]
        cn_font = 'Helvetica'
        for fp in font_paths:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont('CNFont', fp))
                cn_font = 'CNFont'
                break
    except Exception:
        cn_font = 'Helvetica'

    title_style = ParagraphStyle('CNTitle', parent=styles['Title'],
                                  fontName=cn_font, fontSize=16)
    heading_style = ParagraphStyle('CNHeading', parent=styles['Heading2'],
                                    fontName=cn_font, fontSize=12)
    normal_style = ParagraphStyle('CNNormal', parent=styles['Normal'],
                                   fontName=cn_font, fontSize=9)

    elements = []

    # 标题
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10*mm))

    # 总览
    summary = data['summary']
    summary_text = (f"收入: ¥{summary['income']:,.2f}　　"
                    f"支出: ¥{summary['expense']:,.2f}　　"
                    f"结余: ¥{summary['balance']:,.2f}　　"
                    f"共 {data['count']} 笔")
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 8*mm))

    # 分类统计表
    if data['by_category']:
        elements.append(Paragraph('分类统计', heading_style))
        elements.append(Spacer(1, 3*mm))

        cat_data = [['类别', '收入', '支出', '笔数', '占比']]
        total_exp = summary['expense'] or 1
        for cat in data['by_category'][:20]:  # 最多显示20个
            pct = f"{cat['expense'] / total_exp * 100:.1f}%" if total_exp else '0%'
            cat_data.append([
                cat['category'],
                f"¥{cat['income']:,.2f}",
                f"¥{cat['expense']:,.2f}",
                str(cat['count']),
                pct,
            ])

        cat_table = Table(cat_data, colWidths=[80, 70, 70, 40, 50])
        cat_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), cn_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 8*mm))

    # 明细表
    if data['transactions']:
        elements.append(Paragraph('交易明细', heading_style))
        elements.append(Spacer(1, 3*mm))

        tx_data = [['日期', '类型', '金额', '类别', '账户', '商家', '备注']]
        for tx in data['transactions'][:500]:  # 最多500条
            date_short = tx['date'][:16] if tx['date'] else ''
            tx_data.append([
                date_short,
                tx['type'],
                f"¥{tx['amount']:,.2f}",
                tx['category'],
                tx['account'],
                tx['merchant'][:8] if tx['merchant'] else '',
                tx['note'][:10] if tx['note'] else '',
            ])

        tx_table = Table(tx_data,
                         colWidths=[65, 30, 55, 50, 50, 50, 60])
        tx_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), cn_font),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]))
        elements.append(tx_table)

    doc.build(elements)
    return output_path


# ─── 导出预览 ──────────────────────────────────────────────

def get_export_preview(start_date=None, end_date=None, category=None,
                       account=None, type_=None, tag_ids=None):
    """
    获取导出预览信息（不含完整数据，轻量查询）

    返回: { count, date_range, amount_total }
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    where_clauses = ["is_deleted = 0"]
    params = []

    if start_date:
        where_clauses.append("trans_date >= ?")
        params.append(start_date)
    if end_date:
        where_clauses.append("trans_date <= ?")
        params.append(end_date + ' 23:59:59' if len(end_date) == 10 else end_date)
    if category:
        where_clauses.append("(category = ? OR subcategory = ?)")
        params.extend([category, category])
    if account:
        where_clauses.append("account = ?")
        params.append(account)
    if type_:
        where_clauses.append("type = ?")
        params.append(type_)
    if tag_ids:
        placeholders = ','.join('?' * len(tag_ids))
        where_clauses.append(f"""id IN (
            SELECT transaction_id FROM transaction_tags WHERE tag_id IN ({placeholders})
        )""")
        params.extend(tag_ids)

    where_sql = " AND ".join(where_clauses)

    c.execute(f'''SELECT COUNT(*),
                         MIN(trans_date), MAX(trans_date),
                         SUM(CASE WHEN type='收入' THEN amount ELSE 0 END),
                         SUM(CASE WHEN type='支出' THEN amount ELSE 0 END)
                  FROM transactions WHERE {where_sql}''', params)
    row = c.fetchone()
    conn.close()

    count = row[0] or 0
    min_date = (row[1] or '')[:10]
    max_date = (row[2] or '')[:10]
    income = row[3] or 0
    expense = row[4] or 0

    return {
        'count': count,
        'date_range': f"{min_date} ~ {max_date}" if min_date else '无数据',
        'income': income,
        'expense': expense,
        'balance': income - expense,
    }
